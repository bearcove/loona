#!/usr/bin/env -S PYTHONUNBUFFERED=1 FORCE_COLOR=1 uv run

import ctypes
import os
import signal
import subprocess
import sys
import time
import openpyxl
import pprint
from openpyxl.styles import Alignment, Font

from pathlib import Path
from termcolor import colored
import pandas as pd
import tempfile

# Change to the script's directory
os.chdir(os.path.dirname(os.path.abspath(__file__)))

PERF_EVENTS = [
    "cycles",
    "instructions",
    "branches",
    "branch-misses",
    "cache-references",
    "cache-misses",
    "page-faults",
    "syscalls:sys_enter_write",
    "syscalls:sys_enter_writev",
    "syscalls:sys_enter_epoll_wait",
    "syscalls:sys_enter_io_uring_enter"
]
PERF_EVENTS_STRING = ",".join(PERF_EVENTS)

# Define constants for prctl
PR_SET_PDEATHSIG = 1

# Load the libc shared library
libc = ctypes.CDLL("libc.so.6")

def set_pdeathsig():
    """Set the parent death signal to SIGKILL."""
    # Call prctl with PR_SET_PDEATHSIG and SIGKILL
    libc.prctl(PR_SET_PDEATHSIG, signal.SIGKILL)

# Set trap to kill the process group on script exit
def kill_group():
    os.killpg(0, signal.SIGKILL)

def abort_and_cleanup(signum, frame):
    print(colored("\nAborting and cleaning up...", "red"))
    kill_group()
    sys.exit(1)

# Register signal handlers
signal.signal(signal.SIGINT, abort_and_cleanup)
signal.signal(signal.SIGTERM, abort_and_cleanup)

# Create directory if it doesn't exist
Path("/tmp/loona-perfstat").mkdir(parents=True, exist_ok=True)

# Kill older processes
for pidfile in Path("/tmp/loona-perfstat").glob("*.PID"):
    if pidfile.is_file():
        with open(pidfile) as f:
            pid = f.read().strip()
        if pid != str(os.getpid()):
            try:
                os.kill(int(pid), signal.SIGTERM)
            except ProcessLookupError:
                pass
        pidfile.unlink()

# Kill older remote processes
subprocess.run(["ssh", "brat", "pkill -9 h2load"], check=False)

LOONA_DIR = os.path.expanduser("~/bearcove/loona")

# Build the servers
subprocess.run(["cargo", "build", "--release", "--manifest-path", f"{LOONA_DIR}/Cargo.toml", "-F", "tracing/release_max_level_info"], check=True)

# Set protocol, default to h2c
PROTO = os.environ.get("PROTO", "h2c")
os.environ["PROTO"] = PROTO

# Get the default route interface
default_route = subprocess.check_output(["ip", "route", "show", "0.0.0.0/0"]).decode().strip()
default_interface = default_route.split()[4]

# Get the IP address of the default interface
ip_addr_output = subprocess.check_output(["ip", "addr", "show", "dev", default_interface]).decode().strip()

OUR_PUBLIC_IP = None
for line in ip_addr_output.split('\n'):
    if 'inet ' in line:
        OUR_PUBLIC_IP = line.split()[1].split('/')[0]
        break

if not OUR_PUBLIC_IP:
    print(colored("Error: Could not determine our public IP address", "red"))
    sys.exit(1)

print(f"📡 Our public IP address is {OUR_PUBLIC_IP}")

# Declare h2load args based on PROTO
HTTP_OR_HTTPS = "https" if PROTO == "tls" else "http"

# Launch hyper server
hyper_env = os.environ.copy()
hyper_env.update({"ADDR": "0.0.0.0", "PORT": "8001"})
hyper_process = subprocess.Popen([f"{LOONA_DIR}/target/release/httpwg-hyper"], env=hyper_env, preexec_fn=set_pdeathsig)
HYPER_PID = hyper_process.pid
with open("/tmp/loona-perfstat/hyper.PID", "w") as f:
    f.write(str(HYPER_PID))

# Launch loona server
loona_env = os.environ.copy()
loona_env.update({"ADDR": "0.0.0.0", "PORT": "8002"})
loona_process = subprocess.Popen([f"{LOONA_DIR}/target/release/httpwg-loona"], env=loona_env, preexec_fn=set_pdeathsig)
LOONA_PID = loona_process.pid
with open("/tmp/loona-perfstat/loona.PID", "w") as f:
    f.write(str(LOONA_PID))

HYPER_ADDR = f"{HTTP_OR_HTTPS}://{OUR_PUBLIC_IP}:8001"
LOONA_ADDR = f"{HTTP_OR_HTTPS}://{OUR_PUBLIC_IP}:8002"

servers = {
    "hyper": {
        "pid": HYPER_PID,
        "address": HYPER_ADDR
    },
    "loona": {
        "pid": LOONA_PID,
        "address": LOONA_ADDR
    }
}

SERVER = os.environ.get("SERVER")
if SERVER:
    if SERVER in servers:
        servers = {SERVER: servers[SERVER]}
    else:
        print(f"Error: SERVER '{SERVER}' not found in the list of servers.")
        sys.exit(1)

H2LOAD = "/nix/var/nix/profiles/default/bin/h2load"

endpoint = os.environ.get("ENDPOINT", "/repeat-4k-blocks/128")
rps = os.environ.get("RPS", "2")
clients = os.environ.get("CLIENTS", "20")
streams = os.environ.get("STREAMS", "2")
warmup = int(os.environ.get("WARMUP", "5"))
duration = int(os.environ.get("DURATION", "20"))
repeat = int(os.environ.get("REPEAT", "3"))

# Mode can be 'perfstat' or 'samply'
mode = os.environ.get("MODE", "perfstat")

loona_git_sha = subprocess.check_output(['git', 'rev-parse', '--short', 'HEAD'], cwd=os.path.expanduser('~/bearcove/loona')).decode().strip()

benchmark_params = f"RPS={rps}, ENDPOINT={endpoint}, CLIENTS={clients}, STREAMS={streams}, WARMUP={warmup}, DURATION={duration}, REPEAT={repeat}"
print(colored(f"📊 Benchmark parameters: {benchmark_params}", "blue"))

def gen_h2load_cmd(addr: str):
    return [
        H2LOAD,
        "--warm-up-time", str(warmup),
        "--rps", rps,
        "--clients", clients,
        "--max-concurrent-streams", streams,
        "--duration", str(duration),
        f"{addr}{endpoint}",
    ]

def do_samply():
    if len(servers) > 1:
        print(colored("Error: More than one server specified.", "red"))
        print("Please use SERVER=[loona,hyper] to narrow down to a single server.")
        sys.exit(1)

    for server_name, server in servers.items():
        pid = server["pid"]
        address = server["address"]

        print(colored("Warning: Warmup period is not taken into account in samply mode yet.", "yellow"))

        samply_process = subprocess.Popen(["samply", "record", "-p", str(server["pid"])], preexec_fn=set_pdeathsig)
        with open("/tmp/loona-perfstat/samply.PID", "w") as f:
            f.write(str(samply_process.pid))
            h2load_cmd = gen_h2load_cmd(server["address"])
            subprocess.run(["ssh", "brat"] + h2load_cmd, check=True)
        samply_process.send_signal(signal.SIGINT)
        samply_process.wait()

def do_perfstat():
    import pickle
    data_filename = f"/tmp/loona-perfstat/all_data.pkl"

    all_data = None
    if os.environ.get("FROM_DISK") == "1":
        if os.path.exists(data_filename):
            with open(data_filename, 'rb') as f:
                all_data = pickle.load(f)
            print(colored(f"Loaded all_data from {data_filename}", "green"))
        else:
            print(colored(f"Error: File {data_filename} does not exist.", "red"))
            sys.exit(1)

    if all_data is None:
        all_data = {}

        for server_name, server in servers.items():
            pid = server["pid"]
            address = server["address"]

            print(colored(f"🏃 Measuring {server_name} 🏃 (will take {duration} seconds)", "magenta"))
            output_path = tempfile.NamedTemporaryFile(delete=False, suffix='.csv').name

            perf_cmd = [
                "perf", "stat",
                "--event", PERF_EVENTS_STRING,
                "--field-separator", ",",
                "--output", output_path,
                "--pid", str(pid),
                "--delay", str(warmup*1000),
                "--repeat", str(repeat),
                "--",
                "ssh", "brat",
            ] + gen_h2load_cmd(address)
            if PROTO == "tls":
                perf_cmd += ["--alpn-list", "h2"]

            subprocess.run(perf_cmd, preexec_fn=set_pdeathsig, check=True)

            # Read the CSV file, skipping comments and empty lines
            data = []
            with open(output_path, 'r') as f:
                for line in f:
                    if not line.startswith('#') and line.strip():
                        print(f"Using CSV line: {line}")

                        # Split the line by comma and strip whitespace
                        row = [item.strip() for item in line.split(',')]

                        # Ensure we have exactly 4 columns
                        if len(row) >= 4:
                            data.append({
                                'value': row[0],
                                'event': row[2],
                                'stddev': row[3]
                            })

            print("Parsed CSV data:")
            pp = pprint.PrettyPrinter(indent=4)
            pp.pprint(data)

            all_data[server_name] = data

        # Save `all_data` to disk
        import pickle
        with open(data_filename, 'wb') as f:
            pickle.dump(all_data, f)
            print(colored(f"Saved all_data to {data_filename}", "green"))

    print("All data:")
    pp = pprint.PrettyPrinter(indent=4)
    pp.pprint(all_data)

    # Create a new workbook and select the active sheet
    wb = openpyxl.Workbook()

    ws = wb.active

    if ws is None:
        print(colored("Error: Could not find active sheet", "red"))
        sys.exit(1)

    default_font = Font(name='Courier New', size=11)
    bold_font = Font(name='Courier New', size=11, bold=True)

    # Initialize row counter
    current_row = 1

    # Add informative rows
    informative_data = f"date: {time.strftime('%Y-%m-%d')}, loona rev: {loona_git_sha}"
    ws.cell(row=current_row, column=1, value=informative_data)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
    merged_cell = ws.cell(row=current_row, column=1)
    merged_cell.alignment = Alignment(horizontal='left')
    merged_cell.font = bold_font
    current_row += 1

    informative_data = f"{endpoint} over {PROTO}, {repeat} runs each, {rps} reqs/s"
    ws.cell(row=current_row, column=1, value=informative_data)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
    merged_cell = ws.cell(row=current_row, column=1)
    merged_cell.alignment = Alignment(horizontal='left')
    merged_cell.font = bold_font
    current_row += 1

    # Add informative rows
    informative_data = f"{clients} clients with {streams} streams each, {duration}s total duration (including {warmup}s warmup)"
    ws.cell(row=current_row, column=1, value=informative_data)
    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=8)
    merged_cell = ws.cell(row=current_row, column=1)
    merged_cell.alignment = Alignment(horizontal='left')
    merged_cell.font = bold_font
    current_row += 1

    last_informative_row = current_row

    # Add an empty row for spacing
    current_row += 1

    # Add header row
    headers = ["event", "hyper", "", "loona", ""]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = bold_font

    # Merge cells for "hyper" and "loona"
    ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=3)
    ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row, end_column=5)

    # Set alignment for merged cells
    ws.cell(row=current_row, column=2).alignment = Alignment(horizontal='center')
    ws.cell(row=current_row, column=4).alignment = Alignment(horizontal='center')

    # Add subheader row
    current_row += 1
    subheaders = ["", "value", "stddev", "value", "stddev"]
    for col, subheader in enumerate(subheaders, start=1):
        cell = ws.cell(row=current_row, column=col, value=subheader)
        cell.alignment = Alignment(horizontal='center')
        cell.font = Font(bold=True)

    # Center-align the merged cells
    for col in [2, 4]:
        cell = ws.cell(row=current_row-1, column=col)
        cell.alignment = Alignment(horizontal='center')

    # Process and add data rows
    current_row += 1
    for event in PERF_EVENTS:
        event_name = event.split(':')[-1] if ':' in event else event
        ws.cell(row=current_row, column=1, value=event_name)

        for server in ['hyper', 'loona']:
            col = 2 if server == 'hyper' else 4
            if server in all_data:
                server_data = next((item for item in all_data[server] if item['event'] == event), None)
                if server_data:
                    value_cell = ws.cell(row=current_row, column=col, value=float(server_data['value'].replace(',', '')))
                    if event_name in ['cycles', 'instructions', 'branches', 'branch-misses', 'cache-references', 'cache-misses']:
                        value_cell.number_format = '#,##0.0,, "B"'
                    else:
                        value_cell.number_format = '#,##0'

                    stddev = float(server_data['stddev'].rstrip('%')) / 100
                    stddev_cell = ws.cell(row=current_row, column=col+1, value=stddev)
                    stddev_cell.number_format = '0.00%'

        current_row += 1

    # Adjust column widths
    for col in ws.columns:
        max_length = 0
        column = None
        for cell in col[last_informative_row+1:]:  # Skip all informative rows
            if cell.value is not None:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                    column = cell.column_letter
                except TypeError:
                    # Handle cases where len() is not applicable
                    pass
        adjusted_width = (max_length + 2)
        if column is not None:
            ws.column_dimensions[column].width = adjusted_width

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = default_font

    # Save the workbook
    excel_filename = f"/tmp/loona-perfstat/combined_performance.xlsx"
    wb.save(excel_filename)
    print(colored(f"Excel file saved as {excel_filename}", "green"))

try:
    if mode == 'perfstat':
        do_perfstat()
    elif mode == 'samply':
        do_samply()
    else:
        print(colored(f"Error: Unknown mode '{mode}'", "red"))
        print(colored("Known modes:", "yellow"))
        print(colored("- perfstat", "yellow"))
        print(colored("- samply", "yellow"))
        sys.exit(1)
except KeyboardInterrupt:
    print(colored("\nKeyboard interrupt detected. Cleaning up...", "red"))
    abort_and_cleanup(None, None)

kill_group()
